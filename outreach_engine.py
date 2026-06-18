import os
import random
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
# kroa
# tmy

# --- CREDENTIALS & CONFIGURATION ---
# IMPORTANT: Never share this file or upload it to GitHub once your password is here.
SENDER_EMAIL = "try1.naman@gmail.com"  # <--- CHANGE THIS
SENDER_APP_PASSWORD = "fduxqzpjf"  # <--- CHANGE THIS (No spaces)

TRACKER_FILENAME = "D:\\auto apply workflow\\Job_Outreach_Tracker.xlsx"
DAILY_MAX_EMAILS = 50  # Strict safety limit
MIN_DELAY_SECONDS = 5 # 3 minutes minimum pause
MAX_DELAY_SECONDS = 60 # 6 minutes maximum pause

def initialize_tracker(file_path):
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        return wb, sheet
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Outreach Pipeline"
    
    headers = [
        "SNo", "Name", "Email", "Title", "Company", 
        "Status", "Date_Discovered", "Date_Contacted", 
        "Subject_Line", "Follow_Up_Date", "Notes"
    ]
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    sheet.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    wb.save(file_path)
    return wb, sheet

def send_email_via_smtp(recipient_email, subject, body):
    """Connects to Google's SMTP server and sends the email."""
    try:
        # Construct the email structure
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))
        
        # Connect to Gmail SMTP server
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls() # Secure the connection
        server.login(SENDER_EMAIL, SENDER_APP_PASSWORD)
        
        # Send and close
        text = msg.as_string()
        server.sendmail(SENDER_EMAIL, recipient_email, text)
        server.quit()
        return True
    except Exception as e:
        print(f"[ERROR] Failed to send email to {recipient_email}: {e}")
        return False

def generate_personalized_content(candidate_name, candidate_skills, company, recipient_name):
    # Optimal psychological subject line: short, familiar, and specific.
    subject = f"Quick question regarding {company}'s AI team"
    
    first_name = recipient_name.split()[0] if recipient_name else "Team"
    
    body = (
    f"Hi {first_name},\n\n"
    f"I noticed you head up talent focus at {company}.\n\n"
    f"I'm an AI/ML developer specializing in {candidate_skills}. I have built various projects including a self-hosted "
    f"Retrieval-Augmented Generation (RAG) system with local LLMs for secure document intelligence. "
    f"I also work on AI agents and workflow automation systems to solve real-world problems.\n\n"
    f"I believe my experience could align well with {company}'s upcoming AI/ML initiatives. "
    f"Would you be open to a quick 60-second conversation next week to explore potential opportunities?\n\n"
    f"Best regards,\n"
    f"{candidate_name}\n"
    f"GitHub: github.com/aka-naman"
    )
    
    return subject, body

def run_safe_outreach_cycle():
    wb, sheet = initialize_tracker(TRACKER_FILENAME)
    
    if sheet.max_row < 2:
        print("[!] Tracker sheet is empty. Please populate rows with contact info.")
        return
        
    emails_sent_today = 0
    candidate_name = "Naman Sharma"
    candidate_skills = "Python, Deep Learning, NLP, and RAG pipelines" # Shortened for better email flow
    
    header_map = {sheet.cell(row=1, column=i).value: i for i in range(1, sheet.max_row_cols_or_headers() if hasattr(sheet, 'max_row_cols_or_headers') else sheet.max_column + 1)}
    
    for row_idx in range(2, sheet.max_row + 1):
        if emails_sent_today >= DAILY_MAX_EMAILS:
            print(f"\n[Safe Guard] Reached daily limit of {DAILY_MAX_EMAILS} emails. Sleeping until tomorrow.")
            break
            
        status_cell = sheet.cell(row=row_idx, column=header_map.get("Status", 6))
        email_cell = sheet.cell(row=row_idx, column=header_map.get("Email", 3))
        name_cell = sheet.cell(row=row_idx, column=header_map.get("Name", 2))
        company_cell = sheet.cell(row=row_idx, column=header_map.get("Company", 5))
        
        # Only process "Pending" or empty status rows
        if status_cell.value not in [None, "Pending", ""]:
            continue
            
        recipient_email = email_cell.value
        recipient_name = name_cell.value or "Hiring Team"
        company_name = company_cell.value or "your company"
        
        if not recipient_email or "@" not in str(recipient_email):
            status_cell.value = "Skipped (Invalid Email)"
            continue
            
        print(f"\n[+] Processing: {recipient_name} at {company_name} ({recipient_email})")
        
        subject, body = generate_personalized_content(candidate_name, candidate_skills, company_name, recipient_name)
        
        # ACTUALLY SEND THE EMAIL
        print("    -> Sending email via SMTP...")
        success = send_email_via_smtp(recipient_email, subject, body)
        
        if success:
            status_cell.value = "Sent"
            sheet.cell(row=row_idx, column=header_map.get("Date_Contacted", 8)).value = datetime.now().strftime("%Y-%m-%d %H:%M")
            sheet.cell(row=row_idx, column=header_map.get("Subject_Line", 9)).value = subject
            
            follow_up = datetime.now() + timedelta(days=4)
            sheet.cell(row=row_idx, column=header_map.get("Follow_Up_Date", 10)).value = follow_up.strftime("%Y-%m-%d")
            
            emails_sent_today += 1
            wb.save(TRACKER_FILENAME)
            print(f"    [Success] Email sent & logged. Daily count: {emails_sent_today}/{DAILY_MAX_EMAILS}")
            
            if emails_sent_today < DAILY_MAX_EMAILS and row_idx < sheet.max_row:
                sleep_duration = random.randint(MIN_DELAY_SECONDS, MAX_DELAY_SECONDS)
                print(f"    [Anti-Ban] Pausing for {sleep_duration // 60}m {sleep_duration % 60}s...")
                time.sleep(sleep_duration)
        else:
            status_cell.value = "Failed to Send"
            wb.save(TRACKER_FILENAME)

if __name__ == "__main__":
    if SENDER_EMAIL == "your_email@gmail.com":
        print("[!] ERROR: You must update SENDER_EMAIL and SENDER_APP_PASSWORD in the script before running.")
    else:
        run_safe_outreach_cycle()
